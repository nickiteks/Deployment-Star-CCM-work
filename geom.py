import pythoncom
import openpyxl
from win32com.client import Dispatch, gencache
import sys


# Подключение к API7 программы Kompas 3D
def get_kompas_api7():
    module = gencache.EnsureModule("{69AC2981-37C0-4379-84FD-5DD2F3C0A520}", 0, 1, 0)
    api = module.IKompasAPIObject(
        Dispatch("Kompas.Application.7")._oleobj_.QueryInterface(module.IKompasAPIObject.CLSID,
                                                                 pythoncom.IID_IDispatch))
    const = gencache.EnsureModule("{75C9F5D0-B5B8-4526-8681-9903C567D2ED}", 0, 1, 0).constants
    return module, api, const


get_kompas_api7()

module7, api7, const7 = get_kompas_api7()  # Подключаемся к API7
app7 = api7.Application  # Получаем основной интерфейс
app7.Visible = True  # Показываем окно пользователю (если скрыто)
app7.HideMessage = const7.ksHideMessageNo  # Отвечаем НЕТ на любые вопросы программы
print(app7.ApplicationName(FullName=True))  # Печатаем название программы

doc7 = app7.Documents.Open(PathName=sys.argv[1],
                           Visible=True,
                           ReadOnly=True)

#  Подключим константы API Компас
kompas6_constants = gencache.EnsureModule("{75C9F5D0-B5B8-4526-8681-9903C567D2ED}", 0, 1, 0).constants
kompas6_constants_3d = gencache.EnsureModule("{2CAF168C-7961-4B90-9DA2-701419BEEFE3}", 0, 1, 0).constants

#  Подключим описание интерфейсов API5
kompas6_api5_module = gencache.EnsureModule("{0422828C-F174-495E-AC5D-D31014DBBE87}", 0, 1, 0)
kompas_object = kompas6_api5_module.KompasObject(
    Dispatch("Kompas.Application.5")._oleobj_.QueryInterface(kompas6_api5_module.KompasObject.CLSID,
                                                             pythoncom.IID_IDispatch))

#  Подключим описание интерфейсов API7
kompas_api7_module = gencache.EnsureModule("{69AC2981-37C0-4379-84FD-5DD2F3C0A520}", 0, 1, 0)
application = kompas_api7_module.IApplication(
    Dispatch("Kompas.Application.7")._oleobj_.QueryInterface(kompas_api7_module.IApplication.CLSID,
                                                             pythoncom.IID_IDispatch))

Documents = application.Documents
#  Получим активный документ
kompas_document = application.ActiveDocument
kompas_document_3d = kompas_api7_module.IKompasDocument3D(kompas_document)
iDocument3D = kompas_object.ActiveDocument3D()

kPart = iDocument3D.GetPart(kompas6_constants_3d.pTop_Part)

varcoll = kPart.VariableCollection()
varcoll.refresh()

wb_obj = openpyxl.load_workbook(sys.argv[2])

sheet_obj = wb_obj.active

column = sheet_obj.max_column + 1

values = []

for i in range(0, 7):
    values.append(sheet_obj.cell(row=int(sys.argv[3])+1, column=int(sys.argv[4])+1+i).value)

print(values)

list_collms = ['N1','L1','N2','L2','a2','N3','L3']

for i in range(len(values)):
    Variable = varcoll.GetByName(list_collms[i], True, True)
    Variable.value = values[i]

# Перестраиваем модель
kPart.RebuildModel()
# Перерисовываем документ
iDocument3D.RebuildDocument()

savePath = kompas_document.PathName[:-4]
print(savePath)

kompas_document.SaveAs(savePath + f'{sys.argv[3]+1}.stp')
kompas_document.Close(True)